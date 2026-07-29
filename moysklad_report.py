# -*- coding: utf-8 -*-
"""
Ядро отчёта «Лист сборки» из «Мой склад».
Используется и CLI-скриптом, и веб-панелью.

Отчёт строится от ОТГРУЗОК (demand):
  • demand.name                 = номер заказа (совпадает с номером отгрузки);
  • demand.organization         -> «Клиент» (организация);
  • demand.positions.slot       -> «Ячейка» = «<зона> / <ячейка>»;
  • demand.positions.assortment -> товар / артикул / штрихкоды / фото;
  • «Ссылка на этикетку»         -> доп. поле заказа «Этикетка MPsklad».

Штрихкод: из товара (code128); только начинающиеся на «4»; если несколько — все.
"""
import os
import io
import time
import tempfile
from datetime import datetime, timedelta

import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

# ==================== НАСТРОЙКИ ====================
def _load_dotenv():
    """Минимальный загрузчик .env (без зависимостей). Не перезаписывает уже заданные env."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


_load_dotenv()

# Логин/пароль — ТОЛЬКО из окружения (см. .env.example). В коде не хранятся.
LOGIN = os.environ.get("MOYSKLAD_LOGIN")
PASSWORD = os.environ.get("MOYSKLAD_PASSWORD")

ETIKETKA_ATTR_NAME = "Этикетка MPsklad"   # доп. поле заказа со ссылкой на этикетку
WINDOW_HOUR = 13                          # граница суток (13:00)
IMAGE_PX = 80                             # размер картинки в Excel, пикс.
# ===================================================

BASE = "https://api.moysklad.ru/api/remap/1.2"

COLUMNS = ["Ячейка", "Товар", "Фото", "Артикул", "Штрихкод", "Кол-во",
           "Клиент", "№ заказа", "Ссылка на этикетку", "Дата заказа"]


# ---------------- Сессия / запросы ----------------
def get_session():
    if not LOGIN or not PASSWORD:
        raise RuntimeError(
            "Не заданы учётные данные «Мой склад». Укажите MOYSKLAD_LOGIN и "
            "MOYSKLAD_PASSWORD в окружении или в файле .env (см. .env.example).")
    s = requests.Session()
    s.auth = HTTPBasicAuth(LOGIN, PASSWORD)
    s.headers.update({"Accept-Encoding": "gzip", "Content-Type": "application/json"})
    return s


def get_with_retry(session, url, params=None):
    while True:
        resp = session.get(url, params=params)
        if resp.status_code == 429:
            wait = int(resp.headers.get("X-Lognex-Retry-After", 2000)) / 1000 or 2
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp


def _id_from_href(href):
    return (href or "").split("?")[0].rstrip("/").split("/")[-1]


# ---------------- Окно выборки ----------------
def compute_window(now=None, hour=WINDOW_HOUR):
    """Сутки [start, end) с границей hour:00. end = ближайшие прошедшие hour:00."""
    now = now or datetime.now()
    end = now.replace(hour=hour, minute=0, second=0, microsecond=0)
    if now < end:
        end -= timedelta(days=1)
    start = end - timedelta(days=1)
    return start, end


# ---------------- Ячейки склада ----------------
def build_store_slotmap(session, store_id, cache):
    if store_id in cache:
        return cache[store_id]
    zones = {}
    for z in get_with_retry(session, f"{BASE}/entity/store/{store_id}/zones",
                            {"limit": 1000}).json().get("rows", []):
        zones[z.get("id")] = z.get("name", "")
    slotmap = {}
    for sl in get_with_retry(session, f"{BASE}/entity/store/{store_id}/slots",
                             {"limit": 1000}).json().get("rows", []):
        zone = sl.get("zone") or {}
        zid = _id_from_href((zone.get("meta") or {}).get("href")) if zone else None
        zname = zones.get(zid, "")
        sname = sl.get("name", "") or ""
        slotmap[sl.get("id")] = f"{zname} / {sname}" if zname else sname
    cache[store_id] = slotmap
    return slotmap


def extract_cell(session, demand, position, slot_cache):
    slot = position.get("slot") or {}
    slot_id = _id_from_href((slot.get("meta") or {}).get("href")) if slot else None
    if not slot_id:
        return ""
    store_id = _id_from_href(((demand.get("store") or {}).get("meta") or {}).get("href"))
    if not store_id:
        return slot.get("name", "") if isinstance(slot, dict) else ""
    return build_store_slotmap(session, store_id, slot_cache).get(slot_id, "")


# ---------------- Штрихкоды / артикул / этикетка ----------------
def extract_barcodes_4(assortment):
    result = []
    for bc in (assortment.get("barcodes") or []):
        if not isinstance(bc, dict):
            continue
        for value in bc.values():
            if isinstance(value, str) and value.startswith("4") and value not in result:
                result.append(value)
    return result


def extract_etiketka(order):
    for attr in (order.get("attributes") or []):
        if (attr.get("name") or "").strip().lower() == ETIKETKA_ATTR_NAME.lower():
            value = attr.get("value")
            if isinstance(value, dict):
                return value.get("name") or value.get("href") or ""
            return str(value) if value is not None else ""
    return ""


# ---------------- Фото товара ----------------
def download_product_image(session, assortment, img_dir, img_cache):
    pid = assortment.get("id")
    if not pid:
        return None
    if pid in img_cache:
        return img_cache[pid]
    path = None
    images = assortment.get("images") or {}
    href = (images.get("meta") or {}).get("href")
    size = (images.get("meta") or {}).get("size")
    if href and (size is None or size > 0):
        try:
            rows = get_with_retry(session, href, {"limit": 1}).json().get("rows", [])
            if rows:
                dl = ((rows[0].get("miniature") or {}).get("downloadHref")
                      or (rows[0].get("meta") or {}).get("downloadHref"))
                if dl:
                    content = get_with_retry(session, dl).content
                    if content:
                        path = os.path.join(img_dir, f"{pid}.png")
                        with open(path, "wb") as f:
                            f.write(content)
        except Exception:
            path = None
    img_cache[pid] = path
    return path


# ---------------- Загрузка отгрузок ----------------
def fetch_demands(session, start, end):
    start_str = start.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end.strftime("%Y-%m-%d %H:%M:%S")
    url = f"{BASE}/entity/demand"
    params = {
        "limit": 100,
        "offset": 0,
        "order": "moment,desc",
        "expand": "organization,positions.assortment,customerOrder",
        "filter": f"moment>={start_str};moment<{end_str}",
    }
    demands = []
    while True:
        data = get_with_retry(session, url, params).json()
        rows = data.get("rows", [])
        demands.extend(rows)
        meta = data.get("meta", {})
        size, limit, offset = meta.get("size", 0), meta.get("limit", 100), meta.get("offset", 0)
        if not rows or offset + len(rows) >= size:
            break
        params["offset"] = offset + limit
    return demands


def get_positions(session, demand):
    pos_obj = demand.get("positions") or {}
    rows = list(pos_obj.get("rows", []) or [])
    size = (pos_obj.get("meta") or {}).get("size", len(rows))
    if size <= len(rows):
        return rows
    url = f"{BASE}/entity/demand/{demand['id']}/positions"
    params = {"expand": "assortment", "limit": 100, "offset": len(rows)}
    while len(rows) < size:
        chunk = get_with_retry(session, url, params).json().get("rows", [])
        if not chunk:
            break
        rows.extend(chunk)
        params["offset"] += len(chunk)
    return rows


# ---------------- Построение записей ----------------
def _cell_sort_key(cell):
    import re
    m = re.search(r"(\d+)\s*$", str(cell or ""))
    return (0, int(m.group(1))) if m else (1, 0)


def build_records(session, demands, img_dir=None, download_images=True):
    slot_cache, img_cache = {}, {}
    records = []
    total_positions = 0
    for demand in demands:
        order = demand.get("customerOrder") or {}
        number = demand.get("name", "")
        org = (demand.get("organization") or {}).get("name", "")
        etiketka = extract_etiketka(order)
        order_date = order.get("moment") or demand.get("moment") or ""

        positions = get_positions(session, demand)
        total_positions += len(positions)
        if not positions:
            records.append({"Ячейка": "", "Товар": "(позиции отсутствуют)", "Фото": "",
                            "Артикул": "", "Штрихкод": "", "Кол-во": "", "Клиент": org,
                            "№ заказа": number, "Ссылка на этикетку": etiketka,
                            "Дата заказа": order_date, "_img": None, "photo_id": None})
            continue
        for pos in positions:
            if not isinstance(pos, dict):
                continue
            a = pos.get("assortment") or {}
            if not isinstance(a, dict):
                a = {}
            img_path = (download_product_image(session, a, img_dir, img_cache)
                        if (download_images and img_dir) else None)
            records.append({
                "Ячейка": extract_cell(session, demand, pos, slot_cache),
                "Товар": a.get("name", ""),
                "Фото": "",
                "Артикул": a.get("article", ""),
                "Штрихкод": "\n".join(extract_barcodes_4(a)),
                "Кол-во": pos.get("quantity", 0),
                "Клиент": org,
                "№ заказа": number,
                "Ссылка на этикетку": etiketka,
                "Дата заказа": order_date,
                "_img": img_path,
                "photo_id": a.get("id") if img_path else None,
            })
    records.sort(key=lambda r: (_cell_sort_key(r["Ячейка"]), str(r["Товар"])))
    return records, total_positions


# ---------------- Excel ----------------
def build_xlsx_bytes(records, download_images=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист сборки"

    header_fill = PatternFill("solid", fgColor="F5F5F6")
    header_font = Font(name="Arial", bold=True, size=10, color="4A4A4A")
    ws.append(COLUMNS)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    for rec in records:
        ws.append([rec.get(col, "") for col in COLUMNS])

    widths = {"A": 18, "B": 40, "C": 14, "D": 24, "E": 18,
              "F": 8, "G": 30, "H": 16, "I": 55, "J": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wrap = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    if download_images:
        for i, rec in enumerate(records):
            img_path = rec.get("_img")
            excel_row = i + 2
            if img_path and os.path.exists(img_path):
                try:
                    pic = XLImage(img_path)
                    pic.width = IMAGE_PX
                    pic.height = IMAGE_PX
                    ws.add_image(pic, f"C{excel_row}")
                    ws.row_dimensions[excel_row].height = IMAGE_PX * 0.78
                except Exception:
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------- Публичная точка входа ----------------
def generate_report(start=None, end=None, hour=WINDOW_HOUR, download_images=True, img_dir=None):
    """
    Возвращает dict:
      { start, end, records, xlsx, images:{photo_id: path}, stats:{demands, positions, rows} }
    """
    if start is None or end is None:
        start, end = compute_window(hour=hour)
    if download_images and img_dir is None:
        img_dir = tempfile.mkdtemp(prefix="ms_report_img_")

    session = get_session()
    demands = fetch_demands(session, start, end)
    records, total_positions = build_records(session, demands, img_dir, download_images)
    xlsx = build_xlsx_bytes(records, download_images)

    images = {r["photo_id"]: r["_img"] for r in records if r.get("photo_id") and r.get("_img")}
    return {
        "start": start,
        "end": end,
        "records": records,
        "xlsx": xlsx,
        "images": images,
        "stats": {"demands": len(demands), "positions": total_positions, "rows": len(records)},
    }
